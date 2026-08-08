"""Utilitários de leitura e formatação Markdown fiel às células de origem."""
import os, re, unicodedata
from openpyxl import load_workbook

# --------------------------------------------------------------------------
# Identificação oficial do projeto
# Origem: repositório informado pelo proprietário e consultado em 2026-08-07
# https://github.com/edsilas/base-diagnostico-hardware
# --------------------------------------------------------------------------
PROJ_NOME = "Base de Diagnóstico de Hardware"
PROJ_REPO = "base-diagnostico-hardware"
PROJ_AUTOR = "Edsilas"
PROJ_OWNER = "edsilas"
PROJ_URL = "https://github.com/edsilas/base-diagnostico-hardware"
PROJ_DESC = ("Base estruturada de conhecimento para diagnóstico de hardware, com fluxos, "
             "sintomas, códigos de erro, causas e procedimentos de análise e solução.")
PROJ_LICENCA = "MIT"
DOC_VERSAO = "doc-1.3.0"
DOC_DATA = "2026-08-07"

UP = os.environ.get("BDH_FONTES", "./fontes").rstrip("/") + "/"
F_COD = "HW_HARDWARE_CODIGOS_DE_ERROS.xlsx"
F_FLU = "HW_HARDWARE_FLUXO_DIAGNOSTICO.xlsx"


def read(fname):
    wb = load_workbook(UP + fname, data_only=True)
    data = {}
    for ws in wb.worksheets:
        rows = []
        for row in ws.iter_rows(values_only=True):
            if any(c is not None and str(c).strip() != "" for c in row):
                rows.append(["" if c is None else str(c).strip() for c in row])
        # normaliza largura
        w = max(len(r) for r in rows)
        rows = [r + [""] * (w - len(r)) for r in rows]
        data[ws.title] = rows
    wb.close()
    return data


def esc_bs(t):
    """Escapa contrabarra para que caminhos apareçam literalmente no render."""
    return t.replace("\\", "\\\\")


def cell(t):
    """Texto inline (uma linha)."""
    return esc_bs(t.replace("\n", " ").strip())


def tcell(t):
    """Texto para dentro de tabela Markdown."""
    t = esc_bs(t).replace("|", "\\|")
    return re.sub(r"\n+", " <br>", t).strip()


def block(t):
    """Bloco multi-linha preservando a ordem e o texto originais.

    - cada linha não vazia recebe quebra forçada (dois espaços finais);
    - insere linha em branco quando a numeracao reinicia ou termina,
      para que o Markdown nao funda listas distintas.
    """
    if not t.strip():
        return "> Informação não identificada na fonte analisada."
    lines = [l.rstrip() for l in esc_bs(t).split("\n")]
    out, prev = [], None
    for l in lines:
        if not l.strip():
            if out and out[-1] != "":
                out.append("")
            prev = None
            continue
        m = re.match(r"^(\d+)\.\s", l)
        cur = int(m.group(1)) if m else None
        if prev is not None and (cur is None or cur <= prev):
            if out and out[-1] != "":
                out.append("")
        if l.lstrip().startswith("[ ]"):
            l = "- " + l.lstrip()
            # itens de checklist consecutivos formam uma unica lista
            while out and out[-1] == "" and len(out) > 1 and out[-2].startswith("- [ ]"):
                out.pop()
        out.append(l + "  ")
        prev = cur
    txt = "\n".join(out)
    return re.sub(r"\n{3,}", "\n\n", txt).strip()


def field(title, value, level=3):
    """Seção de campo: só emite se houver conteúdo, senão marca a lacuna."""
    h = "#" * level
    if value is None or not str(value).strip():
        return f"{h} {title}\n\n> Informação não identificada na fonte analisada.\n"
    return f"{h} {title}\n\n{block(value)}\n"


def slug(s):
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode()
    s = re.sub(r"[^a-zA-Z0-9]+", "-", s).strip("-").lower()
    return re.sub(r"-{2,}", "-", s)


def gh_anchor(heading_text):
    """Reproduz a regra de âncora do GitHub para um texto de título."""
    s = heading_text.strip().lower()
    s = re.sub(r"[^\w\- ]", "", s, flags=re.UNICODE)
    return s.replace(" ", "-")


# Seções do README usadas na trilha de navegação. As âncoras precisam existir lá.
SECOES = {
    "comecar": ("Comece aqui", "comece-aqui"),
    "diagnosticar": ("Diagnostique", "diagnostique"),
    "resolver": ("Resolva", "resolva"),
    "fechar": ("Feche o atendimento", "feche-o-atendimento"),
    "ferramentas": ("Opere as ferramentas", "opere-as-ferramentas"),
    "referencia": ("Consulte a referência", "consulte-a-referência"),
    "manutencao": ("Manutenção e rastreabilidade", "manutenção-e-rastreabilidade"),
}

HEADER_TPL = """<!-- Gerado a partir de {src}. Não editar manualmente sem atualizar a fonte. -->

{trilha}

# {title}

{resumo}
{aplica}
## Neste documento

<!-- SUMARIO -->

## Contexto

{contexto}

## Escopo

{escopo}

## Fora do escopo

{fora}

## Relação com outros documentos

{rel}

---

"""


def doc_header(title, src, contexto, escopo, fora, rel_list,
               secao="referencia", nivel=0, resumo=None, aplica_se=None):
    """Cabeçalho padrão de todos os documentos.

    secao   — chave em SECOES, define a trilha de navegação até o README.
    nivel   — 0 para docs/*.md, 1 para docs/<subpasta>/*.md.
    resumo  — uma linha explicando o documento, exibida sob o título.
    aplica_se — a que o documento se aplica (linha "Aplica-se a").
    """
    up = "../" * (nivel + 1)
    label, anc = SECOES[secao]
    trilha = (f"[Início]({up}README.md) › [{label}]({up}README.md#{anc}) › **{title}**")
    rel = "\n".join(f"- {r}" for r in rel_list)
    bloco_resumo = f"> {resumo}\n" if resumo else ""
    bloco_aplica = f"\n**Aplica-se a:** {aplica_se}\n" if aplica_se else ""
    return HEADER_TPL.format(title=title, src=src, contexto=contexto, escopo=escopo,
                             fora=fora, rel=rel, trilha=trilha,
                             resumo=bloco_resumo, aplica=bloco_aplica)


FOOTER = """

---

| | |
| --- | --- |
| **Fonte primária deste documento** | {src} |
| **Status de confiança** | {conf} |
| **Última verificação contra a fonte** | {data} |
| **Autoria** | {autor} |
| **Versão da documentação** | `{versao}` |
"""


def doc_footer(src, conf="Confirmado — transcrito das células de origem", data=None,
               proximos=None):
    """Rodapé padrão. `proximos` é uma lista de pares (situação, destino em Markdown)."""
    t = ""
    if proximos:
        t += "\n## Próximos passos\n\n| Se você… | Vá para |\n| --- | --- |\n"
        for cond, destino in proximos:
            t += f"| {cond} | {destino} |\n"
    t += FOOTER.format(src=src, conf=conf, data=data or DOC_DATA,
                       autor=PROJ_AUTOR, versao=DOC_VERSAO)
    return t


# --------------------------------------------------------------------------
# Avisos padronizados (GitHub alerts). Uso consistente em toda a base:
#   nota      — procedência, nível de confiança, observação de leitura
#   dica      — atalho de navegação
#   importante— pré-requisito ou regra que muda o resultado
#   atencao   — risco de erro de diagnóstico ou de perda de tempo
#   perigo    — risco elétrico, perda de dados ou dano a componente
# --------------------------------------------------------------------------
def aviso(tipo, texto):
    tags = {"nota": "NOTE", "dica": "TIP", "importante": "IMPORTANT",
            "atencao": "WARNING", "perigo": "CAUTION"}
    corpo = "\n".join("> " + l if l.strip() else ">" for l in texto.strip().split("\n"))
    return f"> [!{tags[tipo]}]\n{corpo}\n"
